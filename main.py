import requests
import xlrd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import re


class StockDataScraper:
    def __init__(self, url, headers):
        self.url = url
        self.headers = headers

    def fetch_data(self):
        response = requests.get(self.url, headers=self.headers)
        response.encoding = 'GBK'
        if response.status_code == 200:
            return response.text
        else:
            print(f"Error fetching page: HTTP {response.status_code}")
            return None

    def parse_data(self, html):
        soup = BeautifulSoup(html, 'html.parser')
        # 解析股票名
        stock_name = soup.find('h1', string=lambda text: text and text.strip()).get_text(strip=True)
        # 解析股票代码
        stock_code = soup.find('h1', string=lambda text: text and text.strip().isdigit()).get_text(strip=True)
        # 解析市值
        total_market_value_tag = soup.find(string="总市值：")
        if total_market_value_tag:
            total_market_value = total_market_value_tag.find_next(class_="tip f12").text.strip()
        else:
            total_market_value = "没找到总市值"
        # 解析市盈率(静态)
        pe_ratio_static = soup.find(id="jtsyl").text
        # 解析分红率
        dividend_rate = None
        # 解析员工人数
        number_of_employees = None
        # 解析股价
        stock_price = None
        # 解析每股收益
        earnings_per_share_tag = soup.find(string="每股收益：")
        if earnings_per_share_tag:
            earnings_per_share = earnings_per_share_tag.find_next(class_="tip f12").text.strip()
        else:
            earnings_per_share = "没找到每股收益"
        # 解析每股净资产
        net_asset_per_share = soup.find(id="jtsyl").text
        net_asset_per_share_element = soup.find(string="每股净资产：")
        if net_asset_per_share_element:
            net_asset_per_share = net_asset_per_share_element.find_next(class_="tip f12").text.strip()
        else:
            net_asset_per_share = "未找到每股净资产"
        # 解析营业总收入
        total_revenue_element = soup.find('span', class_='hltip f12', string=re.compile(r'营业总收入'))
        if total_revenue_element:
            total_revenue = total_revenue_element.find_next(class_="tip f12").text.strip()
        else:
            total_revenue = "未找到营业总收入"
        # 解析公司亮点
        company_highlight_element = soup.find(string="公司亮点：")
        if company_highlight_element:
            company_highlight = company_highlight_element.find_next(class_="tip f14 fl core-view-text").text.strip()
        else:
            company_highlight = "未找到公司亮点"
        # 解析主营业务
        main_business_element = soup.find('span', class_='hltip f12 fl', string='主营业务：')
        main_business_text = main_business_element.find_next_sibling('span', class_='tip f14 fl main-bussiness-text')
        if main_business_text:
            main_business = main_business_text.get_text(strip=True)
        else:
            main_business = "未找到主营业务"

        translate_map = {'股票名': stock_name,
                         '股票代码': stock_code,
                         '市值': total_market_value,
                         '市盈率': pe_ratio_static,
                         '分红率': dividend_rate,
                         '员工人数': number_of_employees,
                         '股价': stock_price,
                         '每股收益': earnings_per_share,
                         '每股净资产': net_asset_per_share,
                         '营业总收入': total_revenue,
                         '公司亮点': company_highlight,
                         '主营业务': main_business
                         }
        return translate_map

    def extract_text(self, element):
        return element.get_text(strip=True) if element else "未找到"

    def save_to_excel(self, data, ws, wb):
        # ws.append(list(data.keys()))
        ws.append(list(data.values()))
        wb.save("stock_data.xlsx")


def main():
    # 初始化urls列表
    shang_jiao_A_urls, shang_jiao_K_urls, shen_jiao_AK_urls = init_data()
    # 只保留一个excel
    wb = Workbook()
    ws = wb.active
    # 遍历url爬取网站
    for url in shang_jiao_A_urls[:20]:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
        }
        print("开始爬取url:", url)
        scraper = StockDataScraper(url, headers)
        html = scraper.fetch_data()
        if html:
            data = scraper.parse_data(html)
            scraper.save_to_excel(data, ws, wb)

def init_data():
    # 加载 Excel 文件, 两个文件得用不同excel处理库, 不然会报错
    wb_shangjiao_A = xlrd.open_workbook('stock_data/上交所主板A股.xlsx')
    wb_shangjiao_K = xlrd.open_workbook('stock_data/上交所科创板.xlsx')
    wb_shenjiao_AK = load_workbook('stock_data/深交所A股科创股列表.xlsx')

    # 获取第一个工作表
    ws_shangjiao_A = wb_shangjiao_A.sheet_by_index(0)
    ws_shangjiao_K = wb_shangjiao_K.sheet_by_index(0)
    ws_shenjiao_AK = wb_shenjiao_AK.active

    # 创建一个空列表用于存储第一列内容
    shang_jiao_A_stocks = []
    shang_jiao_K_stocks = []
    shen_jiao_AK_stocks = []

    # 遍历第一列的单元格（从第二行开始），并将其值添加到列表中
    for row_index in range(1, ws_shangjiao_A.nrows):
        cell_value = ws_shangjiao_A.cell_value(row_index, 0)  # 第一列索引为0
        shang_jiao_A_stocks.append(cell_value)

    # 遍历第一列的单元格（从第二行开始），并将其值添加到列表中
    for row_index in range(1, ws_shangjiao_K.nrows):
        cell_value = ws_shangjiao_K.cell_value(row_index, 0)  # 第一列索引为0
        shang_jiao_K_stocks.append(cell_value)

    # 遍历第五列的单元格（从第二行开始），并将其值添加到列表中
    for row in ws_shenjiao_AK.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
        shen_jiao_AK_stocks.append(row[0])

    # 使用列表推导式生成url列表
    url_template = 'https://basic.10jqka.com.cn/{}/'
    shang_jiao_A_urls = [url_template.format(code) for code in shang_jiao_A_stocks]
    shang_jiao_K_urls = [url_template.format(code) for code in shang_jiao_K_stocks]
    shen_jiao_AK_urls = [url_template.format(code) for code in shen_jiao_AK_stocks]

    # 返回上交所A股、上交所科创板、深交所A股和科创板
    return shang_jiao_A_urls, shang_jiao_K_urls, shen_jiao_AK_urls

if __name__ == "__main__":
    main()
