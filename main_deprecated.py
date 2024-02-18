import requests
import re
import xlrd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook

# 目标网站的URL
url = 'http://basic.10jqka.com.cn/688311/'

# 模拟浏览器的User-Agent
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

# 发送带有自定义头部的HTTP GET请求
response = requests.get(url, headers=headers)

# 设置响应内容的编码为GBK，以正确解码来自GBK编码页面的内容
response.encoding = 'GBK'

# 确保请求成功
if response.status_code == 200:
    # 使用BeautifulSoup解析HTML，此时传入的HTML文本已经是以GBK编码正确解码的字符串
    soup = BeautifulSoup(response.text, 'html.parser')
    print(soup)

    # 解析股票名
    stock_name = soup.find('h1', string=lambda text: text and text.strip()).get_text(strip=True)
    # 解析股票代码
    stock_code = soup.find('h1', string=lambda text: text and text.strip().isdigit()).get_text(strip=True)
    # 解析市值
    total_market_value_tag = soup.find(string="总市值：")
    if total_market_value_tag:
        total_market_value = total_market_value_tag.find_next(class_="tip f12").text.strip()
    else:
        total_market_value = "没找到总市值"
    # 解析市盈率(静态)
    pe_ratio_static = soup.find(id="jtsyl").text
    # 解析分红率
    dividend_rate = None
    # 解析员工人数
    number_of_employees = None
    # 解析股价
    stock_price = None
    # 解析每股收益
    earnings_per_share_tag = soup.find(string="每股收益：")
    if earnings_per_share_tag:
        earnings_per_share = earnings_per_share_tag.find_next(class_="tip f12").text.strip()
    else:
        earnings_per_share = "没找到每股收益"
    # 解析每股净资产
    net_asset_per_share = soup.find(id="jtsyl").text
    net_asset_per_share_element = soup.find(string="每股净资产：")
    if net_asset_per_share_element:
        net_asset_per_share = net_asset_per_share_element.find_next(class_="tip f12").text.strip()
    else:
        net_asset_per_share = "未找到每股净资产"
    # 解析营业总收入
    total_revenue_element = soup.find('span', class_='hltip f12', string=re.compile(r'营业总收入'))
    if total_revenue_element:
        total_revenue = total_revenue_element.find_next(class_="tip f12").text.strip()
    # 解析公司亮点
    company_highlight_element = soup.find(string="公司亮点：")
    if company_highlight_element:
        company_highlight = company_highlight_element.find_next(class_="tip f14 fl core-view-text").text.strip()
    else:
        company_highlight = "未找到公司亮点"
    # 解析主营业务
    main_business_element = soup.find('span', class_='hltip f12 fl', string='主营业务：')
    main_business_text = main_business_element.find_next_sibling('span', class_='tip f14 fl main-bussiness-text')
    if main_business_text:
        main_business = main_business_text.get_text(strip=True)
    else:
        main_business = "未找到主营业务"

    translate_map = {'股票名': stock_name,
                     '股票代码': stock_code,
                     '市值': total_market_value,
                     '市盈率': pe_ratio_static,
                     '分红率': dividend_rate,
                     '员工人数': number_of_employees,
                     '股价': stock_price,
                     '每股收益': earnings_per_share,
                     '每股净资产': net_asset_per_share,
                     '营业总收入': total_revenue,
                     '公司亮点': company_highlight,
                     '主营业务': main_business
                     }
    for variable_name, variable_value in translate_map.items():
        print(f"{variable_name}: {variable_value}")
    # 创建一个工作簿
    wb = Workbook()
    # 激活默认的工作表
    ws = wb.active
    # 将所有的键写入第一行
    ws.append(list(translate_map.keys()))
    # 将所有的值写入第二行
    ws.append(list(translate_map.values()))
    # 保存 Excel 文件
    wb.save("stock_data.xlsx")

    # 加载 Excel 文件
    wb1 = xlrd.open_workbook('stock_data/上交所主板A股.xlsx')
    wb2 = xlrd.open_workbook('stock_data/上交所科创板.xlsx')
    wb3 = load_workbook('stock_data/深交所A股列表.xlsx')

    # 获取第一个工作表
    ws1 = wb1.sheet_by_index(0)
    ws2 = wb2.sheet_by_index(0)
    ws3 = wb3.active

    # 创建一个空列表用于存储第一列内容
    shang_jiao_A_stocks = []
    shang_jiao_K_stocks = []
    shen_jiao_A_stocks = []

    # 遍历第一列的单元格（从第二行开始），并将其值添加到列表中
    for row_index in range(1, ws1.nrows):
        cell_value = ws1.cell_value(row_index, 0)  # 第一列索引为0
        shang_jiao_A_stocks.append(cell_value)

    # 遍历第一列的单元格（从第二行开始），并将其值添加到列表中
    for row_index in range(1, ws2.nrows):
        cell_value = ws2.cell_value(row_index, 0)  # 第一列索引为0
        shang_jiao_K_stocks.append(cell_value)

    # 遍历第五列的单元格（从第二行开始），并将其值添加到列表中
    for row in ws3.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
        shen_jiao_A_stocks.append(row[0])

    # 打印列表内容
    print(shang_jiao_A_stocks)
    print(shang_jiao_K_stocks)
    print(shen_jiao_A_stocks)
else:
    print(f"Error fetching page: HTTP {response.status_code}")
